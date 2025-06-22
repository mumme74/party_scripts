import csv, json
import openpyxl
from pathlib import Path

import openpyxl.utils
import openpyxl.utils.exceptions
from .helpers import File
from .exceptions import ReadFileException, \
                        ReadFileNotFound, \
                        ReadFileUnhandledFormat, \
                        DataRetrivalError

app_dir = Path(__file__).parent.parent

class DataRow:
    def __init__(self, owner, data=()):
        self.owner = owner
        self.data = tuple(data)

    def __getitem__(self, col):
        try:
            if isinstance(col, str):
                col = self.owner.headers.index(col)
            return self.data[col]
        except ValueError:
            raise DataRetrivalError(f'col: {col} does not exist')
        except IndexError:
            return None
        
    def __iter__(self):
        for vlu in self.data:
            yield vlu

    def __len__(self):
        return len(self.data)
    
    def keys(self):
        for key in self.owner.headers:
            yield key
    
    def values(self):
        for vlu in self.owner.values:
            yield vlu

    def items(self):
        for itm in zip(self.owner.keys(), self.values()):
            yield itm

class Data:
    def __init__(self):
        self.headers:str = ()
        self.rows:DataRow = []
    
    def has_headers(self):
        return len(self.headers) > 0
    
    def keys(self):
        for key in self.headers:
            yield key
    
    def values(self):
        for row in self.rows:
            yield row
    
    def __getitem__(self, row):
        try:
            return self.rows[row]

        except IndexError:
            raise DataRetrivalError(f'row:{row} is invalid')
        
    def __iter__(self):
        for row in self.rows:
            yield row

    def __len__(self):
        return len(self.rows)

def read_data(path: Path, 
              search=[app_dir]) -> list:
    """The root function to read files
    
    Parameter
    ---------
    path: Path
        the path to the file to read

    Returns
    -------
    list: A list of rows

    Raises
    ------
    ReadFileException
    """
    try:
        with File(path, newline='', 
                  encoding='utf8', search=search) as file:
            return _route_to_reader(path, file)
    except (FileNotFoundError):
        raise ReadFileNotFound(path, f"Could not locate {path}")


def _route_to_reader(path: Path, file):
    match path.suffix:
        case '.csv': return _read_csv(path, file, ';')
        case '.tsv': return _read_csv(path, file, '\t')
        case '.json': return _read_json(path, file)
        case '.xlsx': return _read_xlsx(path, file)
        case _: raise ReadFileUnhandledFormat(path, f'{path.suffix} is not handled')

def _read_csv(path, file, delimiter):
    try:
        data = Data()
        reader = csv.reader(file, delimiter=delimiter, 
                            quoting=csv.QUOTE_NONE, strict=True)
        hdrs = [h for h in next(reader)]
        
        # read in all the rows
        for row in reader:
            r = DataRow(data)
            r.data = tuple(v for v in row)
            # add custom columns if not in headers
            for i in range(len(hdrs),len(r.data),1):
                hdrs.append(f'Col{i+1}')

            if len(r.data) < 2:
                raise csv.Error('Bad format, to few columns')
            data.rows.append(r)
        data.headers = tuple(hdrs)
        return data
    except csv.Error as e:
        raise ReadFileException(path, f'{e}')

def _read_json(path, file):
    try:
        obj = json.load(file)
        data = Data()
        hdrs = []
        for row in obj:
            r = DataRow(data)
            d = []
            lsts = sorted(row.items(), 
                              key=lambda x:hdrs.index(x[0]) if x[0] in hdrs else len(hdrs))
            for k,v in lsts:
                if k not in hdrs:
                    hdrs.append(k)
                d.append(v)

            r.data = tuple(d)
            data.rows.append(r)

        data.headers = tuple(hdrs)
        return data
    except json.JSONDecodeError as e:
        raise ReadFileException(path, f'{e}')
    
def _read_xlsx(path, file):
    try:
        frm = openpyxl.load_workbook(path)
        book = frm.active
        data = Data()
        hdrs = [c.internal_value for r in book.iter_rows(1,1) 
                    for c in r if c.internal_value]

        # read in all the rows
        for row in book.iter_rows(2, book.max_row):
            r = DataRow(data)
            r.data = tuple(v.internal_value for v in row)
            # add custom columns if not in headers
            for i in range(len(hdrs),len(r.data),1):
                hdrs.append(f'Col{i+1}')

            if len(r.data) < 2:
                raise csv.Error('Bad format, to few columns')
            data.rows.append(r)
        data.headers = tuple(hdrs)
        return data
    
    except openpyxl.utils.exceptions.InvalidFileException as e:
        raise ReadFileException(path, f'{e}')
